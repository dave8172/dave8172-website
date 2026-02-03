import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def convert_invoice_to_excel(pdf_path, output_path):
    with pdfplumber.open(pdf_path) as pdf:
        wb = Workbook()
        wb.remove(wb.active)
        
        for i, page in enumerate(pdf.pages):
            sheet = wb.create_sheet(title=f"Page {i+1}")
            
            # 1. Extract all text with layout=True
            # This preserves the horizontal spacing between "From" and "Invoice #"
            text = page.extract_text(layout=True)
            
            last_row = 1
            if text:
                for line in text.split('\n'):
                    # Basic check: Stop if we hit the table (usually starts with 'Item' or 'Qty')
                    if "Description" in line or "Item" in line or "Qty" in line:
                        break
                    sheet.cell(row=last_row, column=1, value=line)
                    last_row += 1

            # 2. Extract Table
            table = page.extract_table()
            
            if table:
                # Convert list to DataFrame
                df = pd.DataFrame(table[1:], columns=table[0])
                
                # Add a gap before the table
                last_row += 2
                
                # Use dataframe_to_rows (The fix for your ImportError)
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), last_row):
                    for c_idx, value in enumerate(row, 1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

            # 3. Auto-adjust column widths for better "formatting"
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                sheet.column_dimensions[column].width = max_length + 2

    wb.save(output_path)
    print(f"File saved successfully as: {output_path}")